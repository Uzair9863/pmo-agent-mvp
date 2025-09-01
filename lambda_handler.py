# Week 2: Lambda handler for parsing RAID.xlsx from S3, validating, and saving RAID.json
import json
import boto3
import openpyxl
import os
import tempfile

s3 = boto3.client("s3")

def lambda_handler(event, context):
    bucket_name = "pz-pmo-agent-data"   # your bucket name
    file_key = "RAID_fixed.xlsx"        # Excel file in S3
    output_key = "RAID.json"            # JSON output file in S3

    # Download Excel file from S3 to Lambda's /tmp directory
    tmp_path = os.path.join(tempfile.gettempdir(), "RAID_fixed.xlsx")
    s3.download_file(bucket_name, file_key, tmp_path)

    # Open Excel file with openpyxl
    wb = openpyxl.load_workbook(tmp_path)
    sheet = wb.active

    # ----------------------------
    # Step 1: Read headers
    # ----------------------------
    rows = []
    errors = []
    headers = [cell for cell in next(sheet.iter_rows(values_only=True))]

    # Example required headers
    required_columns = ["Risk", "Impact", "Owner", "Due Date"]

    # Validate headers
    for col in required_columns:
        if col not in headers:
            errors.append(f"Missing required column: {col}")

    if errors:
        return {
            "statusCode": 400,
            "body": json.dumps({
                "message": "Header validation failed",
                "errors": errors
            })
        }

    # ----------------------------
    # Step 2: Read and validate rows
    # ----------------------------
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record = dict(zip(headers, row))

        # Validate required fields
        if not record.get("Risk"):
            errors.append(f"Row {idx}: 'Risk' field is empty")
        if record.get("Impact") not in ["High", "Medium", "Low"]:
            errors.append(f"Row {idx}: Invalid Impact '{record.get('Impact')}'")

        rows.append(record)

    if errors:
        return {
            "statusCode": 400,
            "body": json.dumps({
                "message": "Row validation failed",
                "errors": errors
            })
        }

    # ----------------------------
    # Step 3: Save valid JSON to S3
    # ----------------------------
    json_data = {"rows": rows}

    s3.put_object(
        Bucket=bucket_name,
        Key=output_key,
        Body=json.dumps(json_data),
        ContentType="application/json"
    )

    return {
        "statusCode": 200,
        "body": json.dumps({
            "message": f"Parsed + validated data saved to s3://{bucket_name}/{output_key}",
            "rows": rows
        })
    }



